from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
import time
import random
import os
from openpyxl import Workbook, load_workbook

# -----------------------------
# Setup Excel logging
# -----------------------------
excel_file = "transactions_log.xlsx"

# Create the workbook if it does not exist
if not os.path.exists(excel_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"
    ws.append(["Purchase #", "First Name", "Last Name", "Zip Code", "Items", "Status"])
    wb.save(excel_file)

# Function to log transaction into Excel
def log_transaction(purchase_num, first, last, zipcode, items, status="SUCCESS"):
    wb = load_workbook(excel_file)
    ws = wb.active
    ws.append([purchase_num, first, last, zipcode, ", ".join(items), status])
    wb.save(excel_file)

# -----------------------------
# Selenium automation
# -----------------------------
service = Service(executable_path="chromedriver.exe")
driver = webdriver.Chrome(service=service)

driver.get("https://www.saucedemo.com/")

# Login
username_field = driver.find_element(By.ID, "user-name")
username_field.send_keys("standard_user")
password_field = driver.find_element(By.ID, "password")
password_field.send_keys("secret_sauce")
password_field.send_keys(Keys.RETURN)

time.sleep(2)  # wait for page to load

# Items to add
items_to_add = [
    "add-to-cart-sauce-labs-backpack",
    "add-to-cart-sauce-labs-bike-light",
    "add-to-cart-sauce-labs-bolt-t-shirt"
]

# Example name pools
first_names = ["John", "Emma", "Michael", "Sophia", "David", "Olivia", "James", "Ava"]
last_names = ["Smith", "Johnson", "Williams", "Brown", "Jones", "Garcia", "Miller", "Davis"]

# Number of test runs
num_runs = 20

for i in range(num_runs):
    # Add items to cart
    for item_id in items_to_add:
        try:
            button = driver.find_element(By.ID, item_id)
            button.click()
            print(f"✅ Added {item_id} to cart")
        except Exception as e:
            print(f"⚠️ Could not add {item_id}: {e}")

    # Open cart
    cart_link = driver.find_element(By.CSS_SELECTOR, 'a[data-test="shopping-cart-link"]')
    actions = ActionChains(driver)
    actions.move_to_element(cart_link).click().perform()

    # Checkout
    checkout_btn = driver.find_element(By.CSS_SELECTOR, 'button[data-test="checkout"]')
    checkout_btn.click()

    # Generate random details
    random_first = random.choice(first_names)
    random_last = random.choice(last_names)
    random_zip = str(random.randint(10000, 99999))

    # Input details
    driver.find_element(By.ID, "first-name").send_keys(random_first)
    driver.find_element(By.ID, "last-name").send_keys(random_last)
    driver.find_element(By.ID, "postal-code").send_keys(random_zip)

    # Continue
    driver.find_element(By.ID, "continue").click()

    # Finish
    driver.find_element(By.ID, "finish").click()

    # Log transaction
    log_transaction(i + 1, random_first, random_last, random_zip, items_to_add, "SUCCESS")
    print(f"📒 Logged Purchase #{i+1}")

    # Return to products
    return_btn = driver.find_element(By.ID, "back-to-products")
    return_btn.click()

time.sleep(5)
driver.quit()
